"""
Export Service.

Provides Excel and CSV export functionality for attendance data
using pandas and openpyxl with professional styling.
"""

import pandas as pd
from io import BytesIO
from datetime import date, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models.attendance import Attendance
from models.student import Student


class ExportService:
    """Service class for exporting attendance data to Excel and CSV."""

    @staticmethod
    def _get_attendance_data(start_date, end_date, department=None):
        """Query attendance data and format for export.

        Args:
            start_date: Start date for the query range.
            end_date: End date for the query range.
            department: Optional department filter.

        Returns:
            List of dicts with formatted attendance data.
        """
        query = Attendance.query.join(Student)
        query = query.filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date
        )
        if department:
            query = query.filter(Student.department == department)
        records = query.order_by(
            Attendance.date.desc(), Attendance.time.desc()
        ).all()

        data = []
        for r in records:
            data.append({
                'Date': r.date.strftime('%Y-%m-%d'),
                'Time': r.time.strftime('%H:%M:%S') if r.time else '',
                'Student Name': r.student.name,
                'Roll Number': r.student.roll_number,
                'Department': r.student.department,
                'Status': r.status,
                'Confidence (%)': round(r.confidence * 100, 1) if r.confidence else 'N/A'
            })
        return data

    @staticmethod
    def export_excel(start_date, end_date, department=None):
        """Generate a professionally styled Excel file.

        Args:
            start_date: Start date for the export range.
            end_date: End date for the export range.
            department: Optional department filter.

        Returns:
            BytesIO buffer containing the Excel file.
        """
        data = ExportService._get_attendance_data(
            start_date, end_date, department
        )
        df = pd.DataFrame(data)

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Attendance')

            # Style the workbook
            ws = writer.sheets['Attendance']

            # Header styling
            header_font = Font(
                name='Arial', bold=True, color='FFFFFF', size=11
            )
            header_fill = PatternFill(
                start_color='007AFF', end_color='007AFF', fill_type='solid'
            )
            header_alignment = Alignment(
                horizontal='center', vertical='center'
            )
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col_letter in ws.iter_cols(min_row=1, max_row=1):
                for cell in col_letter:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

            # Data styling
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

            # Auto-fit column widths
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = max_length + 4

        buffer.seek(0)
        return buffer

    @staticmethod
    def export_csv(start_date, end_date, department=None):
        """Generate a CSV file.

        Args:
            start_date: Start date for the export range.
            end_date: End date for the export range.
            department: Optional department filter.

        Returns:
            BytesIO buffer containing the CSV file.
        """
        data = ExportService._get_attendance_data(
            start_date, end_date, department
        )
        df = pd.DataFrame(data)
        buffer = BytesIO()
        df.to_csv(buffer, index=False)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_today():
        """Export today's attendance as an Excel file.

        Returns:
            BytesIO buffer containing the Excel file.
        """
        today = date.today()
        return ExportService.export_excel(today, today)

    @staticmethod
    def export_weekly():
        """Export the current week's attendance as an Excel file.

        Returns:
            BytesIO buffer containing the Excel file.
        """
        today = date.today()
        start = today - timedelta(days=today.weekday())
        return ExportService.export_excel(start, today)

    @staticmethod
    def export_monthly():
        """Export the current month's attendance as an Excel file.

        Returns:
            BytesIO buffer containing the Excel file.
        """
        today = date.today()
        start = today.replace(day=1)
        return ExportService.export_excel(start, today)
